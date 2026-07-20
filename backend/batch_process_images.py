"""
Batch Image Processor -> Excel
===============================

Loops over every image in Test-images/, runs it through the EXISTING
server.py pipeline (unchanged — imported, not modified), and writes one row
per image to an Excel sheet with:

    Column A: Image           -> filename
    Column B: What it sees    -> human-readable list of detected items
    Column C: JSON            -> {"items": [{item, grams, confidence, fdcid}, ...]}

Stage 1 (analyze_meal_image_impl) detects items + grams + confidence.
Stage 2 (lookup_nutrition_impl) is then called per item ONLY to fetch its
USDA fdc_id, since the requested JSON output includes "fdcid" per item.
No other logic from server.py is changed.

Usage:
    python batch_process_images.py
"""

import os
import json
import glob

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Reusing server.py's existing functions exactly as they are — no edits.
from server import analyze_meal_image_impl, lookup_nutrition_impl

IMAGE_DIR = os.path.join(os.path.dirname(__file__), "Test-images")
OUTPUT_XLSX = os.path.join(os.path.dirname(__file__), "meal_image_analysis.xlsx")

IMAGE_EXTENSIONS = ("*.jpg", "*.jpeg", "*.png", "*.webp")


def _list_images(folder: str) -> list[str]:
    paths = []
    for ext in IMAGE_EXTENSIONS:
        paths.extend(glob.glob(os.path.join(folder, ext)))
        paths.extend(glob.glob(os.path.join(folder, ext.upper())))
    return sorted(set(paths))


def _what_it_sees(detected: list[dict]) -> str:
    """Human-readable summary of detected items for column B."""
    lines = []
    for it in detected:
        lines.append(f"{it['item']} (~{it['grams']:g}g, {it['confidence']} confidence)")
    return "\n".join(lines)


def _build_items_json(detected: list[dict]) -> str:
    """Attach fdcid (via existing lookup_nutrition_impl) to each detected
    item and return the {"items": [...]} JSON string in the requested shape."""
    items_with_fdcid = []
    for it in detected:
        nutrition_info = lookup_nutrition_impl(it["item"])
        items_with_fdcid.append({
            "item": it["item"],
            "grams": it["grams"],
            "confidence": it["confidence"],
            "fdcid": str(nutrition_info["fdc_id"]) if nutrition_info["fdc_id"] else None,
        })
    return json.dumps({"items": items_with_fdcid}, indent=2)


def process_folder(image_dir: str, output_path: str) -> None:
    images = _list_images(image_dir)
    if not images:
        print(f"No images found in {image_dir}")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Meal Analysis"

    headers = ["Image", "What it sees", "JSON"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)

    for image_path in images:
        filename = os.path.basename(image_path)
        print(f"\n=== Processing: {filename} ===")

        try:
            detected = analyze_meal_image_impl(image_path=image_path)
            sees_text = _what_it_sees(detected)
            items_json = _build_items_json(detected)
        except Exception as e:
            print(f"[ERROR] Failed on '{filename}': {e}")
            sees_text = f"ERROR: {e}"
            items_json = json.dumps({"items": [], "error": str(e)}, indent=2)

        ws.append([filename, sees_text, items_json])

    # Basic readability formatting
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 70
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(output_path)
    print(f"\nDone. Wrote {len(images)} row(s) to: {output_path}")


if __name__ == "__main__":
    process_folder(IMAGE_DIR, OUTPUT_XLSX)